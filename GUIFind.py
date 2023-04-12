import os
import docx
import openpyxl
import PyPDF2
import warnings
import glob
from tkinter import Tk, filedialog, simpledialog

warnings.filterwarnings("ignore")

# Prompt user for the network share path
root = Tk()
root.withdraw()
network_share_path = filedialog.askdirectory(title="Select Network Share Path")

# Define the search phrases
root = Tk()
root.withdraw()
search_phrases = simpledialog.askstring(title="Search Phrases", prompt="Enter the search phrases separated by commas: ")
search_phrases = [phrase.strip() for phrase in search_phrases.split(',')]

# Define the file extensions to search for
file_extensions = ["docx", "xlsx", "pdf"]

# Open the output file for appending
with open('c:\BGInfo\output2.txt', 'a') as f:
    # Walk through the network share directory and search for files with the specified extensions
    for root, dirs, files in os.walk(network_share_path):
        for file in files:
            file_path = os.path.join(root, file)
            file_ext = file.split(".")[-1].lower()
            if file_ext in file_extensions:
                try:
                    # Search for the phrases in the file contents
                    if file_ext == "docx":
                        doc = docx.Document(file_path)
                        file_contents = "\n".join([para.text for para in doc.paragraphs])
                    elif file_ext == "xlsx":
                        wb = openpyxl.load_workbook(file_path)
                        sheet_names = wb.sheetnames
                        sheet_contents = []
                        for sheet_name in sheet_names:
                            sheet = wb[sheet_name]
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value is not None:
                                        sheet_contents.append(str(cell.value))
                        file_contents = "\n".join(sheet_contents)
                    elif file_ext == "pdf":
                        pdf_file = open(file_path, 'rb')
                        reader = PyPDF2.PdfFileReader(pdf_file)
                        num_pages = reader.getNumPages()
                        page_contents = []
                        for page_num in range(num_pages):
                            page = reader.getPage(page_num)
                            page_text = page.extractText()
                            page_contents.append(page_text)
                        file_contents = "\n".join(page_contents)

                    # Search for the phrases in the file contents
                    for search_phrase in search_phrases:
                        if search_phrase.lower() in file_contents.lower():
                            print(f"Found '{search_phrase}' in file '{file_path}'", file=f)

                except Exception as e:
                    print(f"Error reading '{file_path}': {e}")
