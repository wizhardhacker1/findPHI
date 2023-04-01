import os
import tkinter as tk
import warnings
from tkinter import filedialog

import PyPDF2
import docx
import openpyxl

warnings.filterwarnings("ignore")


def search_files():
    # Define the network share path
    network_share_path = path_entry.get()

    # Define the search phrases
    search_phrases = ["ssn", "mrn", "dob", "password", "birthdate"]

    # Define the file extensions to search for
    file_extensions = ["docx", "xlsx", "pdf"]

    # Open the output file for appending
    with open('output.txt', 'a') as f:
        # Walk through the network share directory and search for files with the specified extensions
        for root, dirs, files in os.walk(network_share_path):
            for file in files:
                file_path = os.path.join(root, file)
                file_ext = file.split(".")[-1].lower()
                if file_ext in file_extensions:
                    try:
                        # Search for the phrases in the file contents
                        if file_ext == "docx":
                            doc = docx.Document(file_path)
                            file_contents = "\n".join([para.text for para in doc.paragraphs])
                        elif file_ext == "xlsx":
                            wb = openpyxl.load_workbook(file_path)
                            sheet_names = wb.sheetnames
                            sheet_contents = []
                            for sheet_name in sheet_names:
                                sheet = wb[sheet_name]
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value is not None:
                                            sheet_contents.append(str(cell.value))
                            file_contents = "\n".join(sheet_contents)
                        elif file_ext == "pdf":
                            pdf_file = open(file_path, 'rb')
                            reader = PyPDF2.PdfFileReader(pdf_file)
                            num_pages = reader.getNumPages()
                            page_contents = []
                            for page_num in range(num_pages):
                                page = reader.getPage(page_num)
                                page_text = page.extractText()
                                page_contents.append(page_text)
                            file_contents = "\n".join(page_contents)

                        # Search for the phrases in the file contents
                        for search_phrase in search_phrases:
                            if search_phrase in file_contents:
                                # Open the output file for writing
                                print(f"Found '{search_phrase}' in file '{file_path}'", file=f)

                    except Exception as e:
                        # Open the output file for writing
                        print(f"Error reading '{file_path}': {e}")


def browse_path():
    folder_selected = filedialog.askdirectory()
    path_entry.delete(0, tk.END)
    path_entry.insert(0, folder_selected)


# Create the main window
root = tk.Tk()
root.title("File Search")

# Create the frame for the path selection
path_frame = tk.Frame(root)
path_frame.pack(side=tk.TOP)

# Create the path label and entry
path_label = tk.Label(path_frame, text="Select folder to search:")
path_label.pack(side=tk.LEFT)
path_entry = tk.Entry(path_frame, width=50)
path_entry.pack(side=tk.LEFT)
browse_button = tk.Button(path_frame, text="Browse", command=browse_path)
browse_button.pack(side=tk.LEFT)

# Create the search button
search_button = tk.Button(root, text="Search", command=search_files)
search_button.pack(side=tk.TOP)

# Run the GUI
root.mainloop()
