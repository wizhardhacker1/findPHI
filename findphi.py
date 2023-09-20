import html
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
import docx
import openpyxl
from PyPDF2 import PdfReader
import datetime
import dateutil.parser as parser

# Function to add specific terms to the list
def add_specific_terms():
    terms_text = entry_search_term.get()
    if terms_text:
        # Split the input text into individual terms
        terms = terms_text.split(',')
        for term in terms:
            term = term.strip()  # Remove leading/trailing spaces
            if term:
                search_terms.append((html.escape(term), "Specific Term"))
        label_status.config(text="Specific Terms Added")

# Function to search for patterns within words using regular expressions
def search_within_words(text, patterns):
    found = []
    for pattern, label in patterns:
        matches = re.finditer(pattern, text, re.IGNORECASE)
        for match in matches:
            found.append((label, match.group(0)))
    return found

# Function to parse a date-like string into a date object
def parse_date(date_string):
    try:
        parsed_date = parser.parse(date_string, fuzzy=True)
        return parsed_date.date()
    except ValueError:
        return None

# Function to search and report
def search_and_report():
    network_share_path = entry_search_path.get()
    output_directory = entry_report_path.get()

    # Disable the Search button and set the message
    button_search.config(state=tk.DISABLED)
    label_status.config(text="Searching, please wait...")

    # Define the search patterns
    search_patterns = [
        (r'\d{3}-\d{2}-\d{4}', "Possible SSN"),  # SSN-like pattern
        (r'\d{2}/\d{2}/\d{4}', "Possible DOB"),  # DOB-like pattern (XX/XX/XXXX or XX-XX-XXXX)
    ]

    found_results = {
        "Possible SSN": [],
        "Possible DOB": [],
        "Possible Password": [],  # Added a section for possible passwords
        "Specific Terms": []  # Added a section for specific terms
    }

    today = datetime.date.today()

    for root, dirs, files in os.walk(network_share_path):
        for file in files:
            file_path = os.path.join(root, file)
            file_ext = file.split(".")[-1].lower()
            if file_ext in ["docx", "xlsx", "pdf", "txt"]:
                try:
                    # Skip temporary Word files
                    if file.startswith("~$"):
                        continue

                    if file_ext == "docx":
                        doc = docx.Document(file_path)
                        file_contents = "\n".join([para.text for para in doc.paragraphs])
                    elif file_ext == "xlsx":
                        wb = openpyxl.load_workbook(file_path)
                        sheet_names = wb.sheetnames
                        sheet_contents = []
                        for sheet_name in sheet_names:
                            sheet = wb[sheet_name]
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value is not None:
                                        sheet_contents.append(str(cell.value))
                        file_contents = "\n".join(sheet_contents)
                    elif file_ext == "pdf":
                        pdf_file = open(file_path, 'rb')
                        reader = PdfReader(pdf_file)

                        # Check if the PDF is encrypted
                        if reader.is_encrypted:
                            # Skip encrypted PDFs automatically
                            continue

                        num_pages = len(reader.pages)
                        page_contents = []
                        for page_num in range(num_pages):
                            page = reader.pages[page_num]
                            page_text = page.extract_text()
                            page_contents.append(page_text)
                        file_contents = "\n".join(page_contents)
                    elif file_ext == "txt":
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as txt_file:
                            file_contents = txt_file.read()

                    # Search for patterns within words
                    found_patterns = search_within_words(file_contents, search_patterns)

                    # Append the findings to the appropriate column
                    for label, matches in found_patterns:
                        if label == "Possible DOB":
                            for dob in matches.split():
                                dob_date = parse_date(dob)
                                if dob_date is not None and dob_date <= today:
                                    found_results[label].append((html.escape(file_path), html.escape(dob)))
                        else:
                            found_results[label].append((html.escape(file_path), html.escape(matches)))

                    # Search for specific terms in file contents
                    for term, label in search_terms:
                        if term.lower() in file_contents.lower():
                            found_results["Specific Terms"].append((html.escape(file_path), term))

                    # Check for potential passwords
                    potential_passwords = find_potential_passwords(file_contents)
                    for password in potential_passwords:
                        found_results["Possible Password"].append((html.escape(file_path), html.escape(password)))

                except Exception as e:
                    # Print detailed error message
                    print(f"Error reading '{file_path}': {e}")

            # Update the progress bar
            progress_value.set((files.index(file) + 1) / len(files) * 100)
            app.update_idletasks()

    # Enable the Search button and set the completion message
    button_search.config(state=tk.NORMAL)
    label_status.config(text="Search completed.")
    progress_value.set(0)  # Clear the progress bar

    # Generate individual HTML reports for each section
    generate_html_reports(output_directory, found_results)

    # Display a message box indicating the search is complete
    messagebox.showinfo("Search Completed", "Search and report generation completed. Results saved to the output directory.")

# Function to generate individual HTML reports for each section
def generate_html_reports(output_directory, found_results):
    os.makedirs(output_directory, exist_ok=True)

    for label, results in found_results.items():
        if results:
            file_name = f"{label.replace(' ', '_').lower()}_report.html"
            file_path = os.path.join(output_directory, file_name)

            with open(file_path, 'w') as f:
                f.write("<html><head><title>PHI Search Results</title></head><body>")
                f.write(f"<h1>{label}</h1>")
                f.write("<table border='1' cellspacing='0' cellpadding='5'><tr><th>File Path</th><th>Matches</th></tr>")
                for file_path, matches in results:
                    f.write(f"<tr><td>{file_path}</td><td>{matches}</td></tr>")
                f.write("</table>")
                f.write("</body></html>")

# Function to find potential passwords using a simple pattern (uppercase, lowercase, and digit)
def find_potential_passwords(text):
    potential_passwords = []
    lines = text.split('\n')  # Split text into lines

    pattern = r'(?=.*[A-Z])(?=.*[a-z])(?=.*\d)(?=.*[@!#$%^&*#])[A-Za-z\d@!#$%^&*#]{5,}'  # Password pattern

    for line in lines:
        words = line.split()  # Split line into words
        for word in words:
            if re.match(pattern, word):  # Check if the word matches the pattern
                potential_passwords.append(word)

    return potential_passwords

# Create the main application window
app = tk.Tk()
app.title("File Search and Reporting")

# Create and configure labels and entry fields
label_search_path = tk.Label(app, text="Enter the path to search:")
entry_search_path = tk.Entry(app)
label_report_path = tk.Label(app, text="Enter the output directory:")
entry_report_path = tk.Entry(app)

# Create and configure buttons
button_search = tk.Button(app, text="Search and Report", command=search_and_report)
button_browse_search = tk.Button(app, text="Browse",
                                 command=lambda: entry_search_path.insert(0, filedialog.askdirectory()))
button_browse_report = tk.Button(app, text="Browse",
                                 command=lambda: entry_report_path.insert(0, filedialog.askdirectory()))

# Create and configure a label for status messages
label_status = tk.Label(app, text="")

# Create a progress bar
progress_value = tk.DoubleVar()
progress_bar = Progressbar(app, mode="determinate", variable=progress_value)

# Create labels, entry fields, and buttons for searching specific terms
label_search_term = tk.Label(app, text="Search for specific terms (comma-separated):")
entry_search_term = tk.Entry(app)
button_search_term = tk.Button(app, text="Add Specific Terms", command=add_specific_terms)

# Organize widgets using the grid layout
label_search_path.grid(row=0, column=0, sticky="e")
entry_search_path.grid(row=0, column=1, columnspan=2, sticky="ew")
button_browse_search.grid(row=0, column=3)
label_report_path.grid(row=1, column=0, sticky="e")
entry_report_path.grid(row=1, column=1, columnspan=2, sticky="ew")
button_browse_report.grid(row=1, column=3)
button_search.grid(row=2, column=0, columnspan=4)
label_status.grid(row=3, column=0, columnspan=4)
progress_bar.grid(row=4, column=0, columnspan=4, sticky="ew")

# Organize widgets for searching specific terms
label_search_term.grid(row=5, column=0, sticky="e")
entry_search_term.grid(row=5, column=1, columnspan=2, sticky="ew")
button_search_term.grid(row=5, column=3)

# Configure column weights for resizing
app.columnconfigure(1, weight=1)
app.columnconfigure(2, weight=1)

# Create a list to store search terms
search_terms = []

# Start the main event loop
app.mainloop()
